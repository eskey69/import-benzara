from __future__ import annotations

import csv
from decimal import Decimal, InvalidOperation
from math import isnan
from pathlib import Path

from openpyxl import load_workbook

from .models import CatalogData, CatalogRow, InventoryData, InventoryIssue


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and isnan(value):
        return ""
    return str(value).strip()


def resolve_header(headers: list[str], expected_name: str) -> str | None:
    expected = expected_name.strip().lower()
    for header in headers:
        if header.strip().lower() == expected:
            return header
    return None


def parse_quantity(raw_value: object) -> int:
    text = normalize_text(raw_value)
    if text == "":
        raise ValueError("Qty is blank")

    try:
        decimal_value = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Qty is not numeric: {text}") from exc

    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f"Qty is not an integer: {text}")

    return int(decimal_value)


def detect_csv_delimiter(path: Path) -> str:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
    if not sample.strip():
        raise ValueError(f"Inventory file is empty: {path}")

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        return ","
    return dialect.delimiter


def load_inventory_csv(path: Path, sku_column: str, qty_column: str, delimiter: str | None) -> InventoryData:
    resolved_delimiter = delimiter or detect_csv_delimiter(path)

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=resolved_delimiter)
        if reader.fieldnames is None:
            raise ValueError(f"Inventory file has no header: {path}")

        field_map = {name.strip(): name for name in reader.fieldnames}
        if sku_column not in field_map or qty_column not in field_map:
            raise ValueError(
                f"Inventory file must contain columns '{sku_column}' and '{qty_column}'. "
                f"Found: {', '.join(reader.fieldnames)}"
            )

        quantities: dict[str, int] = {}
        duplicate_skus: list[InventoryIssue] = []
        invalid_rows: list[InventoryIssue] = []
        conflicting_duplicate_skus: list[InventoryIssue] = []
        row_count = 0

        source_sku = field_map[sku_column]
        source_qty = field_map[qty_column]

        for row_number, row in enumerate(reader, start=2):
            row_count += 1
            sku = normalize_text(row.get(source_sku, ""))
            qty_raw = row.get(source_qty, "")

            if sku == "":
                invalid_rows.append(InventoryIssue(row_number, sku, "SKU is blank"))
                continue

            try:
                qty = parse_quantity(qty_raw)
            except ValueError as exc:
                invalid_rows.append(InventoryIssue(row_number, sku, str(exc)))
                continue

            if sku in quantities:
                duplicate_skus.append(InventoryIssue(row_number, sku, "Duplicate SKU, latest row wins"))
                if quantities[sku] != qty:
                    conflicting_duplicate_skus.append(
                        InventoryIssue(row_number, sku, f"Conflicting duplicate Qty values: {quantities[sku]} vs {qty}")
                    )

            quantities[sku] = qty

    return InventoryData(
        quantities=quantities,
        row_count=row_count,
        duplicate_skus=duplicate_skus,
        invalid_rows=invalid_rows,
        conflicting_duplicate_skus=conflicting_duplicate_skus,
    )


def load_catalog_xlsx(path: Path, sku_column: str, inventory_qty_column: str) -> CatalogData:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]

        row_iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iterator)
        except StopIteration as exc:
            raise ValueError(f"Catalog file is empty: {path}") from exc

        headers = [normalize_text(column) for column in header_row]
        if all(header == "" for header in headers):
            raise ValueError(f"Catalog file has an empty header row: {path}")

        sku_header = resolve_header(headers, sku_column)
        if sku_header is None:
            raise ValueError(f"Catalog file must contain a '{sku_column}' column.")

        inventory_qty_header = resolve_header(headers, inventory_qty_column)

        rows: list[CatalogRow] = []
        for row_number, row in enumerate(row_iterator, start=2):
            values = {}
            for index, header in enumerate(headers):
                if header == "":
                    continue
                raw_value = row[index] if index < len(row) else ""
                values[header] = normalize_text(raw_value)
            rows.append(CatalogRow(row_number=row_number, values=values))

        return CatalogData(
            headers=headers,
            rows=rows,
            sku_header=sku_header,
            inventory_qty_header=inventory_qty_header,
        )
    finally:
        workbook.close()
